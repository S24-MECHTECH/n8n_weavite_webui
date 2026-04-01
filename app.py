import os
import json
import uuid
import zipfile
import tempfile
import shutil
import time
from flask import Flask, request, jsonify
from flask_cors import CORS
import requests

app = Flask(__name__)
CORS(app)

# Config
WEAVIATE_URL = os.environ.get('WEAVIATE_URL', 'http://62.171.136.239:8080')
OLLAMA_URL = os.environ.get('OLLAMA_URL', 'http://62.171.136.239:11434')
CLASS_NAME = os.environ.get('CLASS_NAME', 'Lexware_rag_knowledge')
OLLAMA_MODEL = os.environ.get('OLLAMA_MODEL', 'nomic-embed-text')

# Slow Mode für bessere Stabilität
SLOW_MODE = os.environ.get('SLOW_MODE', 'true').lower() == 'true'
SLOW_DELAY_FILE = float(os.environ.get('SLOW_DELAY_FILE', '0.5'))  # Pause nach jeder Datei
SLOW_DELAY_WEAVIATE = float(os.environ.get('SLOW_DELAY_WEAVIATE', '1.0'))  # Pause nach jedem Weaviate-Request
SLOW_DELAY_OLLAMA = float(os.environ.get('SLOW_DELAY_OLLAMA', '2.0'))  # Pause nach jedem Ollama-Request

TEMP_DIR = '/tmp/uploads'

if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR)

# File type classification for Lexware Buchhalter
# Kategorie 1: Text-Dateien die vektorisiert werden (IMMERS vektorisieren)
VECTORIZE_TYPES = {'pdf', 'txt', 'csv', 'xml', 'json', 'log', 'html', 'rtf', 'docx', 'xlsx'}

# Kategorie 2: Lexware DB-Dateien die extrahiert & vektorisiert werden wenn lesbar
LEXWARE_DB_TYPES = {'lxd', 'lxa', 'lxv', 'f5', 'db', 'dat', 'idx', 'dbf', 'btr', 'cdx', 'fpt'}

# Kategorie 3: Binary-Dateien die NICHT vektorisiert werden (nur Metadata)
NON_VECTORIZE_TYPES = {
    'bak', 'zip', '7z', 'lbu', 'lbk', 'lex',
    'lhd', 'lpd', 'lpe', 'lre', 'lva', 'lza', 'lxs',
    'lfo', 'lfa', 'lwe', 'lwa', 'lwi', 'lwo', 'llg',
    'lmd', 'lna', 'lso', 'elfo', 'elst', 'pfx',
    'mdf', 'ldf', 'ndf', 'trn', 'mt940', 'camt', 'dta', 'sepa',
    'binary', 'lexware_binary', 'unknown'
}

# Mapping für Dateiformate zu Kategorien
FILE_CATEGORIES = {
    # Dokumente (VEKTORISIEREN)
    'pdf': ('Dokument', True), 'txt': ('Dokument', True), 'csv': ('Buchung', True),
    'xml': ('XML', True), 'json': ('JSON', True), 'log': ('Log', True),
    'html': ('HTML', True), 'rtf': ('Dokument', True), 'docx': ('Dokument', True),
    'xlsx': ('Excel', True),
    # Lexware DB-Dateien (VEKTORISIEREN wenn lesbar)
    'lxd': ('Lexware Archiv', True), 'lxa': ('Lexware Archiv', True),
    'lxv': ('Lexware Verz.', True), 'f5': ('F5 Datenbank', True),
    'db': ('Lexware DB', True), 'dat': ('Lexware Datensatz', True),
    'idx': ('Lexware Index', True), 'dbf': ('dBase', True),
    'btr': ('Btrieve', True), 'cdx': ('Index', True), 'fpt': ('FoxPro', True),
    # Binary (NICHT vektorisieren)
    'bak': ('Backup', False), 'zip': ('Archive', False), '7z': ('Archive', False),
    'lbu': ('LXBackup', False), 'lbk': ('LXBackup', False), 'lex': ('Lexware', False),
    'binary': ('Binary', False), 'lexware_binary': ('Lexware Binary', False), 'unknown': ('Unbekannt', False)
}

def classify_file(ext):
    """Classify file type for Lexware Buchhalter"""
    ext = ext.lower()
    if ext in VECTORIZE_TYPES:
        return 'text', True, FILE_CATEGORIES.get(ext, ('Dokument', True))[0]
    elif ext in LEXWARE_DB_TYPES:
        return 'lexware_db', True, FILE_CATEGORIES.get(ext, ('Datenbank', True))[0]
    elif ext in NON_VECTORIZE_TYPES:
        return 'binary', False, FILE_CATEGORIES.get(ext, ('Binary', False))[0]
    else:
        return 'unknown', False, 'Unbekannt'

def get_embedding(text):
    """Get embedding from Ollama"""
    try:
        response = requests.post(
            f'{OLLAMA_URL}/api/embeddings',
            json={'model': OLLAMA_MODEL, 'prompt': text[:8000]},
            timeout=120  # Längeres Timeout für Ollama
        )
        if response.ok:
            # Slow Mode: Pause nach Ollama-Request
            if SLOW_MODE:
                time.sleep(SLOW_DELAY_OLLAMA)
            return response.json().get('embedding')
    except Exception as e:
        print(f"Embedding error: {e}")
    return None

def extract_lexware_db(filepath, ext):
    """Extract text content from Lexware DB/DAT/IDX/F5 files - improved version"""
    try:
        with open(filepath, 'rb') as f:
            raw = f.read()

        file_size = len(raw)
        print(f"  DB-Extraktion: {filepath} ({file_size} bytes, ext={ext})")

        # Try many different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'iso-8859-15',
                     'cp437', 'cp850', 'cp852', 'macroman', 'utf-16', 'utf-16-le', 'utf-16-be']

        for encoding in encodings:
            try:
                text = raw.decode(encoding, errors='ignore')
                # Prüfe ob genug lesbare Zeichen vorhanden sind
                printable = sum(1 for c in text if c.isprintable() or c in '\n\r\t')
                if printable > len(text) * 0.5 and printable > 50:
                    print(f"    ✓ Encoding gefunden: {encoding} ({printable} Zeichen)")
                    return f"[Lexware {ext.upper()} Export | Encoding: {encoding} | Größe: {file_size}]\n{text[:100000]}"
            except Exception as e:
                continue

        # Fallback: Hex-Dump für die ersten 1000 Bytes
        hex_dump = raw[:2000].hex()
        ascii_dump = ''.join(chr(b) if 32 <= b < 127 else '.' for b in raw[:2000])

        # Strukturelle Analyse
        analysis = []
        analysis.append(f"[Lexware {ext.upper()} Datei - Struktur-Analyse]")
        analysis.append(f"Dateigröße: {file_size} bytes")
        analysis.append(f"")
        analysis.append("=== Hex-Dump (erste 1000 Bytes) ===")
        for i in range(0, min(1000, len(raw)), 16):
            chunk = raw[i:i+16]
            hex_str = ' '.join(f'{b:02x}' for b in chunk)
            ascii_str = ''.join(chr(b) if 32 <= b < 127 else '.' for b in chunk)
            analysis.append(f"{i:08x}  {hex_str:<48}  {ascii_str}")

        analysis.append("")
        analysis.append("=== ASCII-Dump ===")
        analysis.append(ascii_dump)

        # Prüfe auf bekannte Patterns
        patterns_found = []
        raw_str = raw[:50000]

        # Suche nach Text-Mustern
        try:
            for encoding in ['utf-8', 'latin-1']:
                try:
                    decoded = raw[:10000].decode(encoding, errors='ignore')
                    # Finde Wörter
                    import re
                    words = re.findall(r'\b[A-Za-z]{3,}\b', decoded)
                    if words:
                        patterns_found.append(f"Gefundene Wörter (Sample): {', '.join(set(words))[:200]}")
                except:
                    pass
        except:
            pass

        if patterns_found:
            analysis.append("")
            analysis.append("=== Pattern-Analyse ===")
            analysis.extend(patterns_found)

        return "\n".join(analysis)

    except Exception as e:
        print(f"DB extraction error: {e}")
        return f"[Lexware {ext} - Extraktionsfehler: {e}]"

def extract_text_from_file(filepath, filename):
    """Extract text based on file extension - returns text only"""
    ext = filename.lower().split('.')[-1]

    # For binary types, return metadata only
    if ext in NON_VECTORIZE_TYPES:
        return f"[Binary Lexware file: {ext}]"

    # For Lexware DB types - try to extract readable content
    if ext in LEXWARE_DB_TYPES:
        text = extract_lexware_db(filepath, ext)
        return text if text else f"[Lexware DB: {filename}]"

    # Text extraction for vectorizable types
    if ext == 'txt':
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

    elif ext == 'pdf':
        try:
            import PyPDF2
            with open(filepath, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ''
                for page in reader.pages:
                    text += page.extract_text() or ''
                if text.strip():
                    return text
        except Exception as e:
            print(f"PDF error: {e}")
        # Try OCR
        try:
            import pytesseract
            from pdf2image import convert_from_path
            images = convert_from_path(filepath)
            text = ''
            for img in images:
                text += pytesseract.image_to_string(img, lang='deu+eng') or ''
            return text
        except Exception as e2:
            print(f"OCR error: {e2}")
        return ''

    elif ext == 'csv':
        import csv
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            return ' '.join([' '.join(row) for row in reader])

    elif ext == 'xml':
        # XML mit besserer Lesbarkeit
        try:
            import xml.etree.ElementTree as ET
            tree = ET.parse(filepath)
            root = tree.get_root()
            # Extrahiere Text aus allen Elementen
            texts = []
            for elem in root.iter():
                if elem.text and elem.text.strip():
                    texts.append(elem.text.strip())
            if texts:
                return '\n'.join(texts)
        except:
            pass
        # Fallback
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            # Entferne XML-Tags für bessere Lesbarkeit
            import re
            text = re.sub(r'<[^>]+>', '', content)
            return text

    elif ext == 'msg':
        # Outlook .msg Dateien - versuche Text zu extrahieren
        try:
            # Versuche mit python-docmsg oder OLE File
            import olefile
            ole = olefile.OleFileIO(filepath)
            # Suche nach RTF oder Plain Text
            for stream in ole.listdir():
                if '0037001F' in stream or '0030001F' in stream:  # PR_BODY / PR_RTF
                    data = ole.openstream(stream).read()
                    try:
                        return data.decode('utf-8', errors='ignore')
                    except:
                        pass
            ole.close()
        except:
            pass
        # Fallback: binäre Datei als Text versuchen
        try:
            with open(filepath, 'rb') as f:
                data = f.read()
                # Versuche verschiedene Encodings
                for enc in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        return data.decode(enc, errors='ignore')
                    except:
                        continue
        except:
            pass
        return f"[Outlook MSG: {filename}]"

    elif ext == 'json':
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            try:
                data = json.load(f)
                return json.dumps(data, indent=2)
            except:
                return f.read()

    elif ext in ['log', 'err', 'trace']:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

    elif ext == 'html':
        from bs4 import BeautifulSoup
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
            return soup.get_text()

    elif ext == 'rtf':
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            import re
            text = f.read()
            text = re.sub(r'\\[a-z]+\d*\s?', '', text)
            text = re.sub(r'[{}]', '', text)
            return text

    elif ext == 'docx':
        try:
            from docx import Document
            doc = Document(filepath)
            return '\n'.join([p.text for p in doc.paragraphs])
        except Exception as e:
            print(f"docx error: {e}")
            return ''

    elif ext == 'xlsx':
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            text = ''
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    text += ' '.join([str(cell.value) if cell.value else '' for cell in row]) + '\n'
            return text
        except Exception as e:
            print(f"xlsx error: {e}")
            return ''

    elif ext in ['mt940', 'camt', 'dta', 'sepa']:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

    else:
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except:
            return f"[Unsupported format: {ext}]"

def get_category(filename):
    """Get category for a file"""
    ext = filename.lower().split('.')[-1]
    _, _, category = classify_file(ext)
    return category

def process_file(filepath, filename):
    """Process single file and return document object"""
    ext = filename.lower().split('.')[-1]
    file_type, should_vectorize, category = classify_file(ext)

    text = extract_text_from_file(filepath, filename)

    if not text or len(text.strip()) < 10:
        text = f"[File: {filename}] [Type: {ext}] [Classified: {category}]"

    doc = {
        'text': text[:50000],
        'quelle': filename,
        'kategorie': category,
        'titel': filename,
        'file_type': ext,
        'vectorize': should_vectorize
    }

    # Get embedding only if should vectorize
    if should_vectorize:
        embedding = get_embedding(text[:8000])
        if embedding:
            doc['_vector'] = embedding

    return doc

def add_to_weaviate(doc, target_class=None):
    """Add document to Weaviate"""
    if target_class is None:
        target_class = CLASS_NAME

    obj = {
        'class': target_class,
        'properties': {
            'text': doc.get('text', ''),
            'quelle': doc.get('quelle', ''),
            'kategorie': doc.get('kategorie', ''),
            'titel': doc.get('titel', '')
        }
    }

    if '_vector' in doc:
        obj['vector'] = doc['_vector']

    try:
        response = requests.post(
            f'{WEAVIATE_URL}/v1/objects',
            headers={'Content-Type': 'application/json'},
            json=obj,
            timeout=60
        )
        # Slow Mode: Pause nach Weaviate-Request
        if SLOW_MODE:
            time.sleep(SLOW_DELAY_WEAVIATE)

        if response.ok:
            return response.json()
        else:
            print(f"Weaviate error: {response.text}")
            return None
    except Exception as e:
        print(f"Weaviate connection error: {e}")
        return None

@app.route('/upload', methods=['POST'])
def upload():
    """Handle file upload"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400

    # Get target class from form data
    target_class = request.form.get('class')
    if not target_class:
        target_class = request.args.get('class', CLASS_NAME)
    if not target_class:
        target_class = CLASS_NAME
    print(f"Upload target class: {target_class}")

    filename = file.filename
    results = []

    # Handle ZIP files
    if filename.lower().endswith('.zip'):
        temp_path = os.path.join(TEMP_DIR, f"{uuid.uuid4()}_{filename}")
        file.save(temp_path)

        extract_dir = tempfile.mkdtemp()
        try:
            try:
                with zipfile.ZipFile(temp_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_dir)
            except NotImplementedError as zip_err:
                # Try with different methods
                try:
                    import subprocess
                    subprocess.run(['unzip', '-o', temp_path, '-d', extract_dir], check=True)
                except:
                    results.append({'file': filename, 'status': 'error', 'error': f'ZIP format not supported: {zip_err}'})
                    return jsonify({'status': 'completed', 'processed': 0, 'results': results})

            # Process all files - SLOW MODE für Stabilität
            all_files = []
            for root, dirs, files in os.walk(extract_dir):
                for fname in files:
                    all_files.append(os.path.join(root, fname))

            total_files = len(all_files)
            print(f"SLOW_MODE: Verarbeite {total_files} Dateien...")

            for i, fpath in enumerate(all_files):
                fname = os.path.basename(fpath)
                print(f"  [{i+1}/{total_files}] Verarbeite: {fname}")

                # Slow Mode: Pause zwischen Dateien
                if SLOW_MODE and i > 0:
                    time.sleep(SLOW_DELAY_FILE)

                try:
                    doc = process_file(fpath, fname)
                    if doc:
                        result = add_to_weaviate(doc, target_class)
                        status = 'success' if result else 'failed'
                        if result:
                            print(f"    ✓ Erfolgreich (ID: {result.get('id', '')[:8]}...)")
                        else:
                            print(f"    ✗ Fehlgeschlagen")
                        results.append({
                            'file': fname,
                            'status': status,
                            'id': result.get('id') if result else None,
                            'vectorized': doc.get('vectorize', False)
                        })
                    else:
                        print(f"    ⚠ Kein Text extrahiert")
                        results.append({'file': fname, 'status': 'skipped', 'error': 'No text extracted'})
                except Exception as e:
                    print(f"    ✗ Fehler: {e}")
                    results.append({'file': fname, 'status': 'error', 'error': str(e)})
        finally:
            try:
                shutil.rmtree(extract_dir)
            except: pass
            try:
                os.remove(temp_path)
            except: pass

    else:
        # Single file
        temp_path = os.path.join(TEMP_DIR, f"{uuid.uuid4()}_{filename}")
        file.save(temp_path)

        try:
            doc = process_file(temp_path, filename)
            if doc:
                result = add_to_weaviate(doc, target_class)
                results.append({
                    'file': filename,
                    'status': 'success' if result else 'failed',
                    'id': result.get('id') if result else None,
                    'vectorized': doc.get('vectorize', False)
                })
            else:
                results.append({'file': filename, 'status': 'error', 'error': 'No text extracted'})
        finally:
            os.remove(temp_path)

    return jsonify({
        'status': 'completed',
        'processed': len(results),
        'results': results
    })

@app.route('/process_folder', methods=['POST'])
def process_folder():
    """Process all files in a folder"""
    data = request.json or {}
    folder_path = data.get('path', '/tmp/uploads')
    target_class = data.get('class', CLASS_NAME)

    if not os.path.exists(folder_path):
        return jsonify({'error': 'Folder not found'}), 400

    results = []
    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    total = len(files)

    for i, fname in enumerate(files):
        fpath = os.path.join(folder_path, fname)
        print(f"  [{i+1}/{total}] {fname}")

        # Slow Mode
        if SLOW_MODE and i > 0:
            time.sleep(SLOW_DELAY_FILE)

        try:
            doc = process_file(fpath, fname)
            if doc:
                result = add_to_weaviate(doc, target_class)
                results.append({
                    'file': fname,
                    'status': 'success' if result else 'failed',
                    'id': result.get('id') if result else None
                })
        except Exception as e:
            results.append({'file': fname, 'status': 'error', 'error': str(e)})

    return jsonify({
        'status': 'completed',
        'processed': len(results),
        'results': results
    })

@app.route('/list_tmp', methods=['GET'])
def list_tmp():
    """List files in TMP folder"""
    folder_path = request.args.get('path', '/home/mechtech/uploads')

    if not os.path.exists(folder_path):
        return jsonify({'error': 'Folder not found', 'path': folder_path}), 400

    try:
        files = []
        for fname in os.listdir(folder_path):
            fpath = os.path.join(folder_path, fname)
            if os.path.isfile(fpath):
                size = os.path.getsize(fpath)
                files.append({
                    'name': fname,
                    'size': size,
                    'path': fpath
                })

        return jsonify({
            'path': folder_path,
            'count': len(files),
            'files': files
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    """Health check"""
    return jsonify({
        'status': 'ok',
        'weaviate': WEAVIATE_URL,
        'ollama': OLLAMA_URL,
        'slow_mode': SLOW_MODE,
        'slow_delay_file': SLOW_DELAY_FILE,
        'slow_delay_weaviate': SLOW_DELAY_WEAVIATE,
        'slow_delay_ollama': SLOW_DELAY_OLLAMA,
        'vectorize_types': list(VECTORIZE_TYPES),
        'lexware_db_types': list(LEXWARE_DB_TYPES),
        'non_vectorize_types': list(NON_VECTORIZE_TYPES),
        'temp_folder': TEMP_DIR
    })

@app.route('/config', methods=['GET', 'POST'])
def config():
    """Get/Set configuration"""
    if request.method == 'POST':
        data = request.json
        global WEAVIATE_URL, OLLAMA_URL, CLASS_NAME, OLLAMA_MODEL
        if 'weaviate_url' in data:
            WEAVIATE_URL = data['weaviate_url']
        if 'ollama_url' in data:
            OLLAMA_URL = data['ollama_url']
        if 'class_name' in data:
            CLASS_NAME = data['class_name']
        if 'ollama_model' in data:
            OLLAMA_MODEL = data['ollama_model']

    return jsonify({
        'weaviate_url': WEAVIATE_URL,
        'ollama_url': OLLAMA_URL,
        'class_name': CLASS_NAME,
        'ollama_model': OLLAMA_MODEL
    })

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
